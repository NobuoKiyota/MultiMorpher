import os
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from pysfx_param_config import PySFXParams
from pysfx_color_config import PySFXColors

class SFXLogger:
    def __init__(self, output_path):
        self.output_path = output_path
        self._init_workbook()
        
    def _init_workbook(self):
        if os.path.exists(self.output_path):
            try:
                self.wb = openpyxl.load_workbook(self.output_path)
                self.ws = self.wb.active
            except:
                self.wb = openpyxl.Workbook()
                self.ws = self.wb.active
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Generation Log"
            # Headers
            headers = ["Score", "File Name"] + [p.name for p in PySFXParams.get_sorted_params()] + ["Date", "GenerationID"]
            self.ws.append(headers)
            self.wb.save(self.output_path)

    def log_entry(self, filename, params, gen_id=""):
        # Reload to ensure concurrent safety (mock)
        # For speed, we keep open, but save often.
        
        row_data = ["", filename] # Score empty
        sorted_params = PySFXParams.get_sorted_params()
        
        for p in sorted_params:
            val = params.get(p.name, "")
            # Round float
            if isinstance(val, float):
                val = round(val, 3)
            row_data.append(val)
            
        row_data.append(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row_data.append(gen_id)
        
        self.ws.append(row_data)
        
        # Style Last Row
        current_row = self.ws.max_row
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for i, cell in enumerate(self.ws[current_row]):
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
            # Param Colors (Offset 2: Score, File)
            if i >= 2 and i < 2 + len(sorted_params):
                p_idx = i - 2
                p = sorted_params[p_idx]
                hex_c = PySFXColors.get_excel_color(p.group)
                if hex_c:
                    cell.fill = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")
    
    def save(self):
        try:
            self.wb.save(self.output_path)
        except Exception as e:
            print(f"Excel Save Error: {e}")

    def update_score(self, filename, score):
        # Reload to find row
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb.active
        
        # Find row by filename (Col B, index 2)
        target_row = None
        for row in ws.iter_rows(min_row=2):
            if row[1].value == filename:
                target_row = row
                break
                
        if target_row:
            target_row[0].value = score
            wb.save(self.output_path)
            return True
        return False
