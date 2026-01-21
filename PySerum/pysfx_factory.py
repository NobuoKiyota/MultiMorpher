import numpy as np
import scipy.io.wavfile
import os
import datetime
import random
from pysfx_engine import PyQuartzEngine, SR, BLOCK_SIZE
from pysfx_engine import PyQuartzEngine, SR, BLOCK_SIZE
from pysfx_factory_logic import GenerationLogic

VERSION = "2.0.0"

class PyQuartzFactory:
    def __init__(self):
        self.engine = PyQuartzEngine()
        self.out_dir = "Output"
        if not os.path.exists(self.out_dir): os.makedirs(self.out_dir)
        
        self.captured_params = {}
        self.recording_params = False

        # Initialize ImageTracer and update PitchType param range
        try:
            from pysfx_image_tracer import ImageTracer
            from pysfx_param_config import PySFXParams
            tracer = ImageTracer()
            count = tracer.get_curve_count()
            for p in PySFXParams.DEFINITIONS:
                if p.name == "PitchType":
                    p.max = 2 + count
                    p.desc = f"Type(0:Flat,1:Lin,2:Exp,3-{2+count}:Img)"
                    break
        except Exception as e:
            print(f"Warning: Failed to init ImageTracer: {e}")

    def _get_param_value(self, config, name):
        """
        Excel設定(config)を優先して値を取得します。
        config[name] = {"min": v, "max": v, "probability": 0-100}
        """
        # 1. Excel Config Check
        if config and name in config:
            node = config[name]
            prob = node.get("probability", 0)
            
            # Probability Check
            if node.get("random", False): prob = 100
            
            if prob > 0 and (prob >= 100 or random.uniform(0, 100) < prob):
                # Apply Random
                v_min = float(node["min"])
                v_max = float(node["max"])
                if v_min > v_max: v_min, v_max = v_max, v_min # Safety
                
                val = random.uniform(v_min, v_max)
                if self.recording_params: self.captured_params[name] = val
                return val
            
            # Use Base Value from Config if available (Probability Failed case)
            if "value" in node and node["value"] is not None:
                val = float(node["value"])
                if self.recording_params: self.captured_params[name] = val
                return val
        
        # 2. Fallback to Hardcoded Default (if not in config)
        from pysfx_param_config import PySFXParams
        default_val = 0.0
        for p in PySFXParams.DEFINITIONS:
            if p.name == name:
                default_val = p.default
                break
        
        # Special case for Boolean defaults
        if default_val is False: default_val = 0.0
        if default_val is True: default_val = 1.0
        
        if self.recording_params:
            self.captured_params[name] = default_val
            
        return float(default_val)

    def run_advanced_batch(self, config, num_files=10, progress_callback=None):
        print(f"--- PyQuartz SFX Factory Production: {num_files} files ---")
        
        for p_idx in range(num_files):
            if progress_callback: progress_callback(p_idx, num_files)
            
            # Reset Log Capture
            self.captured_params = {}
            self.recording_params = True
            
            # 1. 各種パラメータの決定 (Random設定対応)
            duration = self._get_param_value(config, "Duration")
            # NoteRange: Use round() to properly capture random float values (e.g. 60.9 -> 61)
            raw_note = self._get_param_value(config, "NoteRange")
            root_note = int(round(raw_note))
            if self.recording_params:
                 # Override captured with integer used, or keep float? 
                 # Keep float in log for transparency, but debug print here
                 pass
            # print(f"DEBUG: NoteRange raw={raw_note:.2f} -> {root_note}")
            num_voices = int(self._get_param_value(config, "Voices"))
            strum_ms = self._get_param_value(config, "Strum")
            
            # 2. エンジン内部パラメータの同期
            engine_updates = {
                'vol': self._get_param_value(config, "MasterVolume"), 
                'osc_type': int(self._get_param_value(config, "SimpleOscType")),
                'simple_osc_vol': self._get_param_value(config, "SimpleOscVolume"),
                'semi': 0, 
                'cutoff': 20000, 
                'dist': 0,
                'filter_en': True,
                'portamento': self._get_param_value(config, "Portament"),
                'master_pan': self._get_param_value(config, "MasterPan"),
                
                # --- LFO Params ---
                'lfo_p_range': self._get_param_value(config, "LFO_P_Range"),
                'lfo_p_type': int(self._get_param_value(config, "LFO_P_Type")),
                'lfo_p_speed': self._get_param_value(config, "LFO_P_Speed"),
                'lfo_p_shape': self._get_param_value(config, "LFO_P_Shape"),
                
                'lfo_v_range': self._get_param_value(config, "LFO_V_Range"),
                'lfo_v_type': int(self._get_param_value(config, "LFO_V_Type")),
                'lfo_v_speed': self._get_param_value(config, "LFO_V_Speed"),
                'lfo_v_shape': self._get_param_value(config, "LFO_V_Shape"),
                
                'lfo_pan_range': self._get_param_value(config, "LFO_Pan_Range"),
                'lfo_pan_type': int(self._get_param_value(config, "LFO_Pan_Type")),
                'lfo_pan_speed': self._get_param_value(config, "LFO_Pan_Speed"),
                'lfo_pan_shape': self._get_param_value(config, "LFO_Pan_Shape"),
                
                # --- Filter Params ---
                'filter_envamt': self._get_param_value(config, "Filter_EnvAmt"),
                
                'lpf_enable': config["LPF_Enable"]["value"],
                'lpf_cutoff': self._get_param_value(config, "LPF_Cutoff"),
                'lpf_resonance': self._get_param_value(config, "LPF_Resonance"),
                'lpf_autorange': self._get_param_value(config, "LPF_AutoRange"),
                'lpf_resautorange': self._get_param_value(config, "LPF_ResAutoRange"),

                'hpf_enable': config["HPF_Enable"]["value"],
                'hpf_cutoff': self._get_param_value(config, "HPF_Cutoff"),
                'hpf_resonance': self._get_param_value(config, "HPF_Resonance"),
                'hpf_autorange': self._get_param_value(config, "HPF_AutoRange"),
                'hpf_resautorange': self._get_param_value(config, "HPF_ResAutoRange"),
                
                # --- OSC:A Params ---
                'osc_a_vol': self._get_param_value(config, "OSC:A"),
                'osc_a_table': int(self._get_param_value(config, "OSC:A:Table")),
                'osc_a_pos': self._get_param_value(config, "OSC:A:Pos"),
                'osc_a_unison': self._get_param_value(config, "OSC:A:Unison"),
                'osc_a_semi': int(self._get_param_value(config, "OSC:A:Semi")),
                'osc_a_oct': int(self._get_param_value(config, "OSC:A:Oct")),
                # 'osc_a_fine': int(self._get_param_value(config, "OSC:A:Fine")), # Removed from config
                'osc_a_pan': self._get_param_value(config, "OSC:A:Pan"),
                'osc_a_phase': self._get_param_value(config, "OSC:A:Phase"),
                'osc_warpautorange': self._get_param_value(config, "OSC:A:WarpAutoRange"),
                'osc_detuneautorange': self._get_param_value(config, "OSC:A:DetuneAutoRange"),
                # 'osc_a_rand': config["OSC:A:Rand"]["value"], # Boolean REMOVED
                
                # --- OSC:B Params ---
                'osc_b_vol': self._get_param_value(config, "OSC:B"),
                'osc_b_table': int(self._get_param_value(config, "OSC:B:Table")),
                'osc_b_pos': self._get_param_value(config, "OSC:B:Pos"),
                'osc_b_unison': self._get_param_value(config, "OSC:B:Unison"),
                'osc_b_semi': int(self._get_param_value(config, "OSC:B:Semi")),
                'osc_b_oct': int(self._get_param_value(config, "OSC:B:Oct")),
                'osc_b_pan': self._get_param_value(config, "OSC:B:Pan"),
                'osc_b_phase': self._get_param_value(config, "OSC:B:Phase"),
                'osc_b_warpautorange': self._get_param_value(config, "OSC:B:WarpAutoRange"),
                'osc_b_detuneautorange': self._get_param_value(config, "OSC:B:DetuneAutoRange"),
            }
            self.engine.update_params(engine_updates)
            
            # --- Pitch Automation Setup ---
            pitch_range_cents = self._get_param_value(config, "PitchRange")
            # Try/Except for PitchType in case config is old
            try: pitch_type = int(self._get_param_value(config, "PitchType"))
            except: pitch_type = 0
            
            pitch_val = 0
            if pitch_range_cents != 0 and pitch_type != 0:
                pitch_val = int(self._get_param_value(config, "PitchCurve"))
                # Use higher resolution (64 points) for smooth curves
                points = GenerationLogic.get_pitch_curve(pitch_range_cents, 64, curve_val=pitch_val, curve_type=pitch_type)
                range_semis = pitch_range_cents / 100.0
                scaled_points = [(t, v * range_semis) for t, v in points]
                from pysfx_dsp import AutomationLane
                auto = AutomationLane(duration)
                auto.points = scaled_points
                self.engine.automations['pitch'] = auto
            else:
                if 'pitch' in self.engine.automations: del self.engine.automations['pitch']

            # --- Filter Automation Setup ---
            # Helper for Filter automation generation
            def setup_filter_auto(target_key, type_param, range_param_name):
                try: 
                    atype = int(self._get_param_value(config, type_param))
                    arange = self._get_param_value(config, range_param_name)
                    
                    if arange != 0 and atype != 0:
                        # Use default curve/warp of 64 (Linear) since no specific param exists
                        points = GenerationLogic.get_pitch_curve(arange, 64, curve_val=64, curve_type=atype)
                        # Points are (t, 0.0-1.0). Range scaling handled in Engine (val * range).
                        # So we just pass raw 0.0-1.0 value? 
                        # Engine does: (auto_val * auto_depth).
                        # So automation should return 0.0 to 1.0.
                        # Wait, get_pitch_curve returns (t, val). Val depends on logic.
                        # Logic returns 0.0-1.0 if 'Image' or 'Linear'.
                        # So we assign directly.
                        from pysfx_dsp import AutomationLane
                        auto = AutomationLane(duration)
                        auto.points = points
                        self.engine.automations[target_key] = auto
                    else:
                         if target_key in self.engine.automations: del self.engine.automations[target_key]
                except Exception as e:
                    # print(f"Auto setup err {target_key}: {e}")
                    pass

            setup_filter_auto('lpf_auto', "LPF_AutoType", "LPF_AutoRange")
            setup_filter_auto('lpf_res_auto', "LPF_ResAutoType", "LPF_ResAutoRange")
            setup_filter_auto('hpf_auto', "HPF_AutoType", "HPF_AutoRange")
            setup_filter_auto('hpf_res_auto', "HPF_ResAutoType", "HPF_ResAutoRange")
            setup_filter_auto('warp_auto', "OSC:A:WarpAutoType", "OSC:A:WarpAutoRange")
            setup_filter_auto('detune_auto', "OSC:A:DetuneAutoType", "OSC:A:DetuneAutoRange")
            setup_filter_auto('warp_auto_b', "OSC:B:WarpAutoType", "OSC:B:WarpAutoRange")
            setup_filter_auto('detune_auto_b', "OSC:B:DetuneAutoType", "OSC:B:DetuneAutoRange")

            # 3. ADSRの設定
            a = self._get_param_value(config, "AmpAttack") / 1000.0
            d = self._get_param_value(config, "AmpDecay") / 1000.0
            s = self._get_param_value(config, "AmpSustain") / 127.0
            r = self._get_param_value(config, "AmpRelease") / 1000.0
            self.engine.set_adsr(a, d, s, r)

            # Filter ADSR
            fa = self._get_param_value(config, "Filter_Attack") / 1000.0
            fd = self._get_param_value(config, "Filter_Decay") / 1000.0
            fs = self._get_param_value(config, "Filter_Sustain") # 0-1
            fr = self._get_param_value(config, "Filter_Release") / 1000.0
            self.engine.set_filter_adsr(fa, fd, fs, fr)

            # 4. ノート生成
            is_chord = config["Chord"]["value"]
            notes, chord_name = GenerationLogic.get_chord_notes(root_note, num_voices, is_chord)
            
            # --- Voice & Detune Parameters ---
            vol_root = self._get_param_value(config, "RouteVoiceVolume")
            vol_multi = self._get_param_value(config, "MultiVoiceVolume")
            
            detune_count = int(self._get_param_value(config, "DetuneVoice"))
            detune_range_cents = self._get_param_value(config, "DetuneRange")
            detune_vol_ratio = self._get_param_value(config, "DetuneVolume")
            
            # 5. レンダリング
            # Phase 1: Hold (Gate Open)
            hold_samples = int(duration * SR)
            strum_samples = int((strum_ms / 1000.0) * SR)
            total_hold_blocks = int(np.ceil((hold_samples + (strum_samples * len(notes))) / BLOCK_SIZE))
            
            audio_buffer = []
            current_sample_global = 0
            
            for b in range(total_hold_blocks):
                block_start = current_sample_global
                for idx, n in enumerate(notes):
                    onset = idx * strum_samples
                    # Note On Timing
                    if block_start <= onset < block_start + BLOCK_SIZE:
                        # Determine Volume
                        velocity = vol_root if idx == 0 else vol_multi
                        
                        # 1. Main Voice
                        main_pan = self._get_param_value(config, "Pan")
                        self.engine.note_on(n, velocity=velocity, detune=0.0, pan=main_pan)
                        
                        # 2. Detune Voices
                        if detune_count > 0:
                            det_vol = velocity * detune_vol_ratio
                            for _ in range(detune_count):
                                d_cent = random.uniform(-detune_range_cents, detune_range_cents)
                                d_semi = d_cent / 100.0
                                d_pan = self._get_param_value(config, "Pan")
                                self.engine.note_on(n, velocity=det_vol, detune=d_semi, pan=d_pan)

                raw_data = self.engine.generate_block()
                audio_buffer.append(raw_data)
                current_sample_global += BLOCK_SIZE
            
            # Phase 2: Release & Tail Generation
            for n in notes: self.engine.note_off(n)
            
            # --- Smart Tail Logic ---
            # Generate enough silence to capture Reverb/Delay tails.
            # User requested ~5s timeout.
            # We generate dry silence, which FX chain will fill with tails.
            
            tail_sec = 5.0 
            max_tail_blocks = int(tail_sec * SR / BLOCK_SIZE)
            
            # We just generate the full buffer. Trimming happens POST-FX.
            # Premature optimization (breaking on dry silence) kills Reverb tails.
            
            for b_tail in range(max_tail_blocks):
                raw_data = self.engine.generate_block()
                audio_buffer.append(raw_data)

            
            # 6. 保存と後処理
            full_audio = np.concatenate(audio_buffer)

            # Reshape Interleaved (..., 2) for Stereo
            if len(full_audio) % 2 != 0:
                full_audio = full_audio[:len(full_audio)-1]
            
            stereo_audio = full_audio.reshape(-1, 2)
            
            # --- POST-PROCESSING FX CHAIN ---
            from pysfx_effects import EffectsProcessor
            
            # 1. Distortion
            dist_gain = self._get_param_value(config, "DistortionGain")
            dist_tone = self._get_param_value(config, "DistortionFeed")
            dist_wet = self._get_param_value(config, "DistortionWet")
            if dist_wet > 0:
                stereo_audio[:, 0] = EffectsProcessor.apply_distortion(stereo_audio[:, 0], dist_gain, dist_tone, dist_wet)
                stereo_audio[:, 1] = EffectsProcessor.apply_distortion(stereo_audio[:, 1], dist_gain, dist_tone, dist_wet)

            # 2. Phaser
            ph_depth = self._get_param_value(config, "PhaserDepth")
            ph_speed = self._get_param_value(config, "PhaserSpeed")
            ph_wet = self._get_param_value(config, "PhaserWet")
            if ph_wet > 0:
                stereo_audio[:, 0] = EffectsProcessor.apply_phaser(stereo_audio[:, 0], ph_depth, ph_speed, ph_wet)
                stereo_audio[:, 1] = EffectsProcessor.apply_phaser(stereo_audio[:, 1], ph_depth, ph_speed, ph_wet)

            # 3. Delay
            dly_time = self._get_param_value(config, "DelayTime")
            dly_fb = self._get_param_value(config, "DelayFeedback")
            dly_wet = self._get_param_value(config, "DelayWet")
            if dly_wet > 0:
                stereo_audio[:, 0] = EffectsProcessor.apply_delay(stereo_audio[:, 0], dly_time, dly_fb, dly_wet)
                stereo_audio[:, 1] = EffectsProcessor.apply_delay(stereo_audio[:, 1], dly_time, dly_fb, dly_wet)

            # 4. Reverb
            rv_time = self._get_param_value(config, "ReverbTime")
            rv_spread = self._get_param_value(config, "ReverbSpread")
            rv_wet = self._get_param_value(config, "ReverbWet")
            if rv_wet > 0:
                stereo_audio[:, 0] = EffectsProcessor.apply_reverb(stereo_audio[:, 0], rv_time, rv_spread, rv_wet)
                stereo_audio[:, 1] = EffectsProcessor.apply_reverb(stereo_audio[:, 1], rv_time, rv_spread, rv_wet)
                
            # 5. Spread
            sp_range = self._get_param_value(config, "SpreadRange")
            sp_density = self._get_param_value(config, "SpreadDensity")
            sp_wet = self._get_param_value(config, "SpreadWet")

            # --- Smart Silence Trimming (Backward) ---
            # Trim trailing silence down to the actual signal end.
            # Threshold: 1e-5 (-100dB)
            trim_thresh = 1e-5
            
            # Check magnitude (Mono Sum or Max)
            mag = np.abs(stereo_audio)
            is_loud = np.max(mag, axis=1) > trim_thresh
            
            if np.any(is_loud):
                # Find last index where signal exists
                last_idx = np.where(is_loud)[0][-1]
                
                # Add tiny buffer (e.g. 100ms) for fade out safety? 
                # User said "Trim that 500ms" of silence. 
                # If we detected 500ms of silence, we cut it.
                # Here we just cut *all* silence after last signal.
                # Let's add 100ms safety to avoid sharp cuts on Reverb tails.
                safety_samples = int(0.1 * SR)
                cut_point = min(len(stereo_audio), last_idx + safety_samples)
                
                stereo_audio = stereo_audio[:cut_point]
            else:
                # All silence?
                pass

            # Flatten for Normalize logic need Update?
            
            # --- Fade Out Processing ---
            fade_out_val = self._get_param_value(config, "FadeOutTime")
            fade_out_curve = self._get_param_value(config, "FadeOutCurve")
            
            if fade_out_val > 0.001:
                total_samples = len(stereo_audio)
                # Fade Length: relative to total length. Max (1.0) = 50% length.
                fade_len = int(total_samples * fade_out_val * 0.5) 
                
                if fade_len > 0:
                    start_fade_idx = total_samples - fade_len
                    # Linear 0 to 1
                    lin = np.linspace(0.0, 1.0, fade_len)
                    
                    # Mapping: power = 10.0 ** ((0.5 - curve) * 2) 
                    power = 10.0 ** ((0.5 - fade_out_curve) * 2.0)
                    # Use (1-x) base for fade out
                    fade_curve = (1.0 - lin) ** power
                    
                    # Apply
                    # stereo_audio is stereo (N, 2)
                    fade_curve = fade_curve[:, np.newaxis]
                    stereo_audio[start_fade_idx:] *= fade_curve

            do_normalize = config["Normalize"]["value"]
            peak = np.max(np.abs(stereo_audio))
            
            if do_normalize and peak > 0:
                stereo_audio = (stereo_audio / peak) * 0.95
            elif peak > 1.0: 
                stereo_audio = (stereo_audio / peak) * 0.99
            
            timestamp = datetime.datetime.now().strftime("%H%M%S%f")[:7]
            
            note_str = GenerationLogic.get_note_name(root_note)
            filename = f"Quartz_{note_str}_{chord_name}_{timestamp}.wav"
            filepath = os.path.join(self.out_dir, filename)
            
            scipy.io.wavfile.write(filepath, SR, stereo_audio.astype(np.float32))
            
            # --- Excel Logging (XLSX) ---
            from pysfx_param_config import PySFXParams
            from pysfx_color_config import PySFXColors
            import openpyxl
            from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
            
            xlsx_path = os.path.join(self.out_dir, "generation_log.xlsx")
            
            if os.path.exists(xlsx_path):
                try:
                    wb = openpyxl.load_workbook(xlsx_path)
                    ws = wb.active
                except:
                    # Fallback if corrupted
                    wb = openpyxl.Workbook()
                    ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Generation Log"
                # Headers: Score, File Name, Params..., Date, Version
                headers = ["Score", "File Name"] + [p.name for p in PySFXParams.get_sorted_params()] + ["Date", "Version"]
                ws.append(headers)
            
            # Check Header (Migration for Score column)
            if ws['A1'].value != "Score":
                ws.insert_cols(1)
                ws['A1'] = "Score"
            
            # Construct Row
            # Score is empty by default (User fills 1-10)
            row_data = ["", filename]
            sorted_params = PySFXParams.get_sorted_params()
            
            for p in sorted_params:
                # 1. Try Captured (Prioritize actual used value)
                if p.name in self.captured_params:
                    val = self.captured_params[p.name]
                # 2. Try Config (Fallback)
                elif p.name in config:
                    node = config[p.name]
                    if isinstance(node, dict):
                        val = node.get('value', "")
                    else:
                        val = node # Should not happen if strictly typed
                else:
                    val = ""
                
                # Round floats
                if isinstance(val, (float, np.floating)):
                     val = round(val, 2)
                
                # Convert Bool to INT/String for easier reading? Or keep Bool.
                row_data.append(val)
                
            row_data.append(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            row_data.append(VERSION)
            ws.append(row_data)
            
            # Styling
            current_row = ws.max_row
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for i, cell in enumerate(ws[current_row]):
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                # Colorize based on column
                # Col 0(A)=Score, 1(B)=File, 2(C)=ParamStart
                if i >= 2 and i < len(row_data) - 1:
                    param_idx = i - 2
                    if param_idx < len(sorted_params):
                        p = sorted_params[param_idx]
                        hex_c = PySFXColors.get_excel_color(p.group)
                        if hex_c:
                            cell.fill = PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")
                            
            try:
                wb.save(xlsx_path)
                print(f"[{p_idx+1}/{num_files}] {filename} -> Logged")
            except Exception as e:
                print(f"[{p_idx+1}/{num_files}] Log Error: {e}")

    # --- New Features ---

    def get_random_config(self):
        from pysfx_param_config import PySFXParams
        config = {}
        for p in PySFXParams.get_sorted_params():
            config[p.name] = {
                "value": p.default, 
                "min": p.min if p.min is not None else 0,
                "max": p.max if p.max is not None else 1,
                "random": True 
            }
            if p.min is None: # Boolean
                config[p.name]['value'] = random.choice([True, False])
                config[p.name]['random'] = False
        return config

    def load_favorites(self, min_score=8):
        """Load entries with Score >= min_score"""
        xlsx_path = os.path.join(self.out_dir, "generation_log.xlsx")
        if not os.path.exists(xlsx_path): return []
        
        import openpyxl
        from pysfx_param_config import PySFXParams
        
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        
        # Identify Header Map
        headers = [c.value for c in ws[1]]
        
        if "Score" not in headers: return []
        score_idx = headers.index("Score")
        
        favorites = []
        sorted_params = PySFXParams.get_sorted_params()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            score = row[score_idx]
            if not isinstance(score, (int, float)): continue
            if score < min_score: continue
            
            entry = {}
            # Parse params
            for p in sorted_params:
                if p.name in headers:
                    p_idx = headers.index(p.name)
                    val = row[p_idx]
                    entry[p.name] = val
            favorites.append((score, entry))
        
        # Sort by score desc
        favorites.sort(key=lambda x: x[0], reverse=True)
        return [x[1] for x in favorites]

    def get_similar_config(self, entry):
        from pysfx_param_config import PySFXParams
        config = {}
        variance = 0.1 # 10%
        
        for p in PySFXParams.get_sorted_params():
            base_val = entry.get(p.name, p.default)
            if base_val is None or base_val == "": base_val = p.default

            if p.min is None:
                config[p.name] = {"value": base_val, "random": False}
                continue

            total_range = (p.max - p.min)
            delta = total_range * variance
            v_min = max(p.min, base_val - delta)
            v_max = min(p.max, base_val + delta)
            
            config[p.name] = {
                "value": base_val,
                "min": v_min,
                "max": v_max,
                "random": True
            }
        return config

    def get_hybrid_config(self, entry_a, entry_b):
        from pysfx_param_config import PySFXParams
        config = {}
        
        files_discrete = ["OSC:A:Table", "OSC:B:Table", "NoteRange", "PitchType", "LFO_P_Type", "LFO_V_Type", "LFO_Pan_Type", "Chord"]
        
        for p in PySFXParams.get_sorted_params():
            val_a = entry_a.get(p.name, p.default)
            val_b = entry_b.get(p.name, p.default)
            
            if val_a is None: val_a = p.default
            if val_b is None: val_b = p.default
            
            final_val = val_a
            
            # Boolean
            if p.min is None:
                final_val = random.choice([val_a, val_b])
            # Discrete
            elif any(k in p.name for k in files_discrete) or isinstance(p.default, int):
                final_val = random.choice([val_a, val_b])
            else:
                # Average
                final_val = (val_a + val_b) / 2.0
                
            config[p.name] = {"value": final_val, "random": False}
            
        return config