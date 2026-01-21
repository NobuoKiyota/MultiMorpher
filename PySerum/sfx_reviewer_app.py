import customtkinter as ctk
import os
import shutil
import openpyxl
import winsound
import threading
import json
from tkinter import filedialog, ttk
import tkinter as tk

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("blue")

class SFXReviewerApp(ctk.CTk):
    def __init__(self, start_path=None):
        super().__init__()
        self.title("SFX Factory Reviewer 2.0")
        self.geometry("1100x850")
        
        self.current_folder = ""
        self.excel_path = ""
        self.wb = None
        self.ws = None
        self.data_map = {} # filename -> row_index
        self.current_file = None
        
        self.auto_advance = True
        self.tag_vars = {} # tag_name -> BooleanVar
        self.layout_config = self.load_layout_config()
        self.quick_tags = self.load_tag_config()
        
        # Apply Geometry from Config
        geo = self.layout_config.get("window_geometry", "1100x850")
        self.geometry(geo)
        
        self._init_ui()

        
        if start_path and os.path.exists(start_path):
             self.load_batch(start_path)
        else:
            # Try load last state
            try:
                if os.path.exists("last_state.json"):
                    with open("last_state.json", "r") as f:
                        state = json.load(f)
                        last = state.get("last_reviewer_path")
                        if last and os.path.exists(last):
                            self.load_batch(last)
            except: pass
        
    def load_tag_config(self):
        default_tags = ["Noisy", "Clean", "Click", "Hum", "Distortion", "LowEnd", "HighFreq", "Metallic", "Organic", "Long", "Short", "Loopable"]
        try:
            if os.path.exists("tagger_config.json"):
                with open("tagger_config.json", "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return data.get("quick_tags", default_tags)
        except: pass
        return default_tags

    def load_layout_config(self):
        default_layout = {
            "window_geometry": "1100x850",
            "tags_expand": False,
            "tags_columns": 4,
            "params_height": 0, # 0 means default/flexible
            "params_show_mode": "full" # full or summary
        }
        try:
            if os.path.exists("tagger_config.json"):
                with open("tagger_config.json", "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return data.get("layout", default_layout)
        except: pass
        return default_layout

    def _init_ui(self):
        # Layout: Left (List 300px), Right (Detail)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # --- Left Panel ---
        fr_left = ctk.CTkFrame(self, width=300)
        fr_left.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        fr_left.grid_rowconfigure(1, weight=1)
        
        ctk.CTkButton(fr_left, text="Load Batch Folder", command=self.browse_folder).grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        # Language Toggle
        self.lang_mode = "EN" # EN or JP
        self.btn_lang = ctk.CTkButton(fr_left, text="Lang: EN", width=60, fg_color="#555555", command=self.toggle_language)
        self.btn_lang.grid(row=3, column=1, padx=5, pady=20, sticky="e")
        
        ctk.CTkButton(fr_left, text="☁ Sync to Cloud", command=self.on_sync, fg_color="#009688", hover_color="#00796B").grid(row=3, column=0, padx=10, pady=20, sticky="ew")
        
        # Treeview style list
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#2b2b2b", fieldbackground="#2b2b2b", foreground="white", rowheight=24)
        style.map("Treeview", background=[("selected", "#1f538d")])
        
        self.tree = ttk.Treeview(fr_left, columns=("score", "name"), show="headings")
        self.tree.heading("score", text="Scr", anchor="center")
        self.tree.column("score", width=40, anchor="center")
        self.tree.heading("name", text="Filename", anchor="w")
        self.tree.column("name", width=220)
        
        self.tree.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        
        sb = ttk.Scrollbar(fr_left, orient="vertical", command=self.tree.yview)
        sb.grid(row=1, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=sb.set)
        
        # Filter / Stats
        self.lbl_stats = ctk.CTkLabel(fr_left, text="Total: 0 | Rated: 0")
        self.lbl_stats.grid(row=2, column=0, pady=5)
        
        # --- Right Panel ---
        fr_right = ctk.CTkFrame(self)
        fr_right.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        
        # Row 0: Header
        self.lbl_filename = ctk.CTkLabel(fr_right, text="No File Selected", font=("Arial", 20, "bold"), anchor="w")
        self.lbl_filename.grid(row=0, column=0, sticky="w", padx=20, pady=(10, 0))
        
        self.lbl_info = ctk.CTkLabel(fr_right, text="...", font=("Arial", 12), text_color="gray", anchor="w")
        self.lbl_info.grid(row=1, column=0, sticky="w", padx=20, pady=(0, 10))
        
        # Row 2: Tagging Area
        fr_tags = ctk.CTkFrame(fr_right, fg_color="#222222")
        fr_tags.grid(row=2, column=0, sticky="ew", padx=20, pady=5)
        ctk.CTkLabel(fr_tags, text="Tags:", font=("Arial", 12, "bold")).pack(anchor="w", padx=10, pady=5)
        
        # Checkbox Grid
        fr_checks = ctk.CTkFrame(fr_tags, fg_color="transparent")
        fr_checks.pack(fill="x", padx=10, pady=5)
        
        cols = self.layout_config.get("tags_columns", 4)
        
        # Mapping for language switching
        self.chk_widgets = [] # (checkbox_obj, en_text, jp_text)
        
        for i, tag_data in enumerate(self.quick_tags):
            # Parse Format: "Tag" or ["Tag", "Info"]
            if isinstance(tag_data, list) or isinstance(tag_data, tuple):
                tag_val = str(tag_data[0])
                tag_disp_en = tag_val
                tag_disp_jp = str(tag_data[1]) if len(tag_data) > 1 else tag_val
            else:
                tag_val = str(tag_data)
                tag_disp_en = tag_val
                tag_disp_jp = tag_val
            
            var = ctk.BooleanVar()
            self.tag_vars[tag_val] = var
            chk = ctk.CTkCheckBox(fr_checks, text=tag_disp_en, variable=var, font=("Arial", 11))
            chk.grid(row=i//cols, column=i%cols, sticky="w", padx=5, pady=5)
            
            self.chk_widgets.append((chk, tag_disp_en, tag_disp_jp))
            
        # Custom Entry
        fr_cust = ctk.CTkFrame(fr_tags, fg_color="transparent")
        fr_cust.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(fr_cust, text="Custom:", width=60).pack(side="left")
        self.ent_custom_tags = ctk.CTkEntry(fr_cust, placeholder_text="Comma separated tags...")
        self.ent_custom_tags.pack(side="left", fill="x", expand=True)
        # Bind FocusOut to auto-save or at least ensure entry doesn't lose data
        # self.ent_custom_tags.bind("<FocusOut>", lambda e: self.save_current_tags())


        # Configurable Layout Logic
        tags_expand = self.layout_config.get("tags_expand", False)
        
        if tags_expand:
            # Expand Tags Area (Row 2), Compressing Params
            fr_right.grid_rowconfigure(2, weight=1) 
            fr_right.grid_rowconfigure(3, weight=0)
        else:
            # Default: Share or prioritize params
            fr_right.grid_rowconfigure(3, weight=1)

        # Row 3: Params (Expandable or Fixed Small)
        p_height = self.layout_config.get("params_height", 0)
        
        # If showing summary, we might not need a ScrollableFrame if it's just 1 line, 
        # but keeping it consistent is safer or swapping to simple Frame.
        if self.layout_config.get("params_show_mode") == "summary":
            # Use simple frame for summary to save space and remove scrollbar
            self.fr_params = ctk.CTkFrame(fr_right)
            # Create a label inside for the summary text
            self.lbl_param_summary = ctk.CTkLabel(self.fr_params, text="Params: ...", text_color="gray")
            self.lbl_param_summary.pack(pady=10, padx=10, anchor="w")
            
            if p_height > 0:
                self.fr_params.configure(height=p_height)
            self.fr_params.grid(row=3, column=0, sticky="ew", padx=20, pady=10)
            
        else:
            # Full List Mode
            self.fr_params = ctk.CTkScrollableFrame(fr_right, label_text="Parameters")
            if p_height > 0:
                self.fr_params.configure(height=p_height)
            self.fr_params.grid(row=3, column=0, sticky="nsew", padx=20, pady=10)
        
        # Row 4: Controls Zone (Bottom)
        fr_ctrl = ctk.CTkFrame(fr_right, height=120, fg_color="#1a1a1a")
        fr_ctrl.grid(row=4, column=0, sticky="ew", padx=20, pady=20)
        
        # Current Score
        self.lbl_current_score = ctk.CTkLabel(fr_ctrl, text="-", font=("Arial", 64, "bold"), width=100)
        self.lbl_current_score.pack(side="left", padx=30)
        
        fr_btns = ctk.CTkFrame(fr_ctrl, fg_color="transparent")
        fr_btns.pack(side="left", fill="y", padx=10)
        
        ctk.CTkLabel(fr_btns, text="Numpad 1-9 to Score").pack(anchor="w")
        self.chk_auto = ctk.CTkCheckBox(fr_btns, text="Auto Advance", command=self.toggle_auto)
        self.chk_auto.select()
        self.chk_auto.pack(anchor="w", pady=5)
        
        ctk.CTkButton(fr_btns, text="PLAY (.)", command=self.play_audio, fg_color="green", width=120).pack(pady=10)
        
        # --- Binds ---
        self.bind("<Key>", self.on_key)
        self.bind(".", lambda e: self.play_audio())
        
    def toggle_auto(self):
        self.auto_advance = self.chk_auto.get()
        
    def toggle_language(self):
        if self.lang_mode == "EN":
            self.lang_mode = "JP"
            self.btn_lang.configure(text="Lang: JP")
        else:
            self.lang_mode = "EN"
            self.btn_lang.configure(text="Lang: EN")
            
        # Update Checkboxes
        for chk, en, jp in self.chk_widgets:
            chk.configure(text=jp if self.lang_mode == "JP" else en)
            
    def browse_folder(self):
        p = filedialog.askdirectory()
        if p:
            self.load_batch(p)
            
    def load_batch(self, folder):
        self.current_folder = folder
        
        # Save state
        try:
            state = {}
            if os.path.exists("last_state.json"):
                with open("last_state.json", "r") as f:
                    state = json.load(f)
            state["last_reviewer_path"] = folder
            with open("last_state.json", "w") as f:
                json.dump(state, f)
        except: pass
        
        # Look for Excel
        xls = os.path.join(folder, "final_manifest.xlsx")
        if not os.path.exists(xls):
             xls = os.path.join(folder, "generation_log.xlsx")
             
        if not os.path.exists(xls):
            print("No Excel log found!")
            return
            
        self.excel_path = xls
        self.wb = openpyxl.load_workbook(xls)
        self.ws = self.wb.active
        
        # Headers
        self.headers = [c.value for c in self.ws[1]]
        
        # Identify Columns
        self.idx_score = 0 # Default A
        self.idx_name = 1  # Default B
        self.idx_tags = -1
        self.idx_version = -1
        
        for i, h in enumerate(self.headers):
            if h == "Score": self.idx_score = i
            elif h == "File Name": self.idx_name = i
            elif h == "Tags": self.idx_tags = i
            elif h == "Version": self.idx_version = i
            
        # Ensure Headers exist if missing (Update Excel later)
        if self.idx_tags == -1:
            self.idx_tags = len(self.headers)
            self.headers.append("Tags")
            self.ws.cell(row=1, column=self.idx_tags+1, value="Tags")
            
        # Clear Tree
        for i in self.tree.get_children(): self.tree.delete(i)
        self.data_map = {}
        
        rated_c = 0
        
        for idx, row in enumerate(self.ws.iter_rows(min_row=2)):
            # Row index is idx+2 (Excel is 1-based, min_row=2)
            # cell object is row[i]
            
            # Safe access
            def get_v(i): return row[i].value if i < len(row) else None
            
            score = get_v(self.idx_score)
            fname = get_v(self.idx_name)
            
            if not fname: continue
            
            # Map filename
            self.data_map[fname] = {
                "row_obj": row,
                "vals": [c.value for c in row],
                "path": self.find_file(fname)
            }
            
            # Add to Tree
            disp_score = str(score) if score else "-"
            item_id = self.tree.insert("", "end", values=(disp_score, fname))
            self.data_map[fname]["item_id"] = item_id
            
            if score:
                rated_c += 1
                try:
                    sC = int(score)
                    if sC >= 8: self.tree.item(item_id, tags=("high",))
                    elif sC <= 3: self.tree.item(item_id, tags=("low",))
                    else: self.tree.item(item_id, tags=("mid",))
                except: pass
        
        self.tree.tag_configure("high", foreground="#4caf50")
        self.tree.tag_configure("low", foreground="#f44336")
        self.tree.tag_configure("mid", foreground="#FFC107") # Amber
        
        self.lbl_stats.configure(text=f"Total: {len(self.data_map)} | Rated: {rated_c}")
        
    def find_file(self, fname):
        # 1. Root
        p = os.path.join(self.current_folder, fname)
        if os.path.exists(p): return p
        
        # 2. Score Folders (1-9)
        for i in range(1, 10):
            p = os.path.join(self.current_folder, f"Score_{i}", fname)
            if os.path.exists(p): return p
            
        # 3. Legacy High/Low
        p = os.path.join(self.current_folder, "HighScore", fname)
        if os.path.exists(p): return p
        p = os.path.join(self.current_folder, "LowScore", fname)
        if os.path.exists(p): return p
        
        return None

    def on_select(self, event):
        # Save previous before loading new
        if self.current_file:
            self.save_current_tags()
            
        sel = self.tree.selection()
        if not sel: return
        item = self.tree.item(sel[0])
        fname = item['values'][1]
        self.load_file_details(fname)
        
        if self.auto_advance:
             self.play_audio()

    def save_current_tags(self):
        if not self.current_file: return
        fname = self.current_file
        data = self.data_map.get(fname)
        if not data: return
        
        row = data["row_obj"]
        
        # Gather Tags
        active_tags = [k for k, v in self.tag_vars.items() if v.get()]
        custom_txt = self.ent_custom_tags.get().strip()
        if custom_txt:
            # Avoid dupes
            existing = set(active_tags)
            for t in custom_txt.split(","):
                ts = t.strip()
                if ts and ts not in existing:
                    active_tags.append(ts)
            
        final_tag_str = ", ".join(active_tags)
        
        # Check if changed (optimization)
        row_idx = row[0].row
        current_cell_val = self.ws.cell(row=row_idx, column=self.idx_tags+1).value
        
        if str(current_cell_val) != final_tag_str:
            self.ws.cell(row=row_idx, column=self.idx_tags+1, value=final_tag_str)
            try:
                self.wb.save(self.excel_path)
            except: pass # Silent fail on quick switch


    def load_file_details(self, fname):
        data = self.data_map.get(fname)
        if not data: return
        
        self.current_file = fname
        self.lbl_filename.configure(text=fname)
        
        # Path & Version
        ver = "Unknown"
        row = data["row_obj"]
        
        # Get Version from row
        if self.idx_version != -1 and self.idx_version < len(row):
            v_val = row[self.idx_version].value
            if v_val: ver = str(v_val)
            
        self.lbl_info.configure(text=f"Path: {data['path'] if data['path'] else 'MISSING'}  |  Engine Ver: {ver}")
        
        # Score
        sc = row[self.idx_score].value
        self.lbl_current_score.configure(text=str(sc) if sc else "-")
        self.apply_score_color(sc)
             
        # Tags Restore
        # Reset bits
        for var in self.tag_vars.values(): var.set(False)
        self.ent_custom_tags.delete(0, "end")
        
        if self.idx_tags != -1 and self.idx_tags < len(row):
            tag_str = row[self.idx_tags].value
            if tag_str:
                tags = [t.strip() for t in str(tag_str).split(",")]
                customs = []
                for t in tags:
                    if t in self.tag_vars:
                        self.tag_vars[t].set(True)
                    elif t:
                        customs.append(t)
                if customs:
                    self.ent_custom_tags.insert(0, ", ".join(customs))
        
        # Params Display Logic
        show_mode = self.layout_config.get("params_show_mode", "full")
        
        vals = [c.value for c in row]
        
        if show_mode == "summary":
            # Count params
            total_p = 0
            used_p = 0
            params_detail = []
            
            for i, h in enumerate(self.headers):
                if h in ("Score", "File Name", "Tags", "Version", "Date"): continue
                total_p += 1
                if i < len(vals) and vals[i] not in (None, ""):
                    used_p += 1
                    # Optional: collect names of used params for tooltip or minimal display?
                    # For now just count.
            
            summary_text = f"Parameters Used: {used_p} / {total_p}"
            if hasattr(self, "lbl_param_summary"):
                self.lbl_param_summary.configure(text=summary_text)
                
        else:
            # Full List
            for w in self.fr_params.winfo_children(): w.destroy()
            
            for i, h in enumerate(self.headers):
                # Skip non-params
                if h in ("Score", "File Name", "Tags", "Version", "Date"): continue
                # Also params might be None
                if i >= len(vals): val = ""
                else: val = vals[i]
                
                row_fr = ctk.CTkFrame(self.fr_params, fg_color="transparent")
                row_fr.pack(fill="x", pady=1)
                
                ctk.CTkLabel(row_fr, text=str(h)[:25], width=180, anchor="w", text_color="#aaaaaa").pack(side="left")
                ctk.CTkLabel(row_fr, text=str(val), anchor="w", font=("Consolas", 12)).pack(side="left")

    def apply_score_color(self, sc):
        if not sc:
            self.lbl_current_score.configure(text_color="white")
            return
        try:
            s_int = int(sc)
            if s_int >= 8: 
                self.lbl_current_score.configure(text_color="#4caf50")
            elif s_int <= 3:
                self.lbl_current_score.configure(text_color="#f44336")
            else:
                self.lbl_current_score.configure(text_color="#FFC107")
        except:
             self.lbl_current_score.configure(text_color="white")

    def play_audio(self):
        if not self.current_file: return
        path = self.data_map[self.current_file]["path"]
        if path and os.path.exists(path):
            winsound.PlaySound(path, winsound.SND_ASYNC)

    def on_key(self, event):
        if not self.current_file: return
        if event.char.isdigit():
            val = int(event.char)
            if 0 <= val <= 9:
                if val == 0: val = 10 # Map 0 to 10
                self.set_score(val)
        
    def set_score(self, score):
        if not self.current_file: return
        fname = self.current_file
        data = self.data_map[fname]
        row = data["row_obj"]
        
        # 1. Gather Tags (Already saved by save_current_tags implicitly if we call it, but let's be explicit)
        self.save_current_tags() # Ensure tags are flushed
        
        # Re-fetch because save_current_tags writes to Excel but mainly we want to ensure latest state
        # Note: save_current_tags saves tags. set_score saves score.
        pass
        
        # 2. Update Excel Row (Score)

        
        # 2. Update Excel Row
        # Score
        row[self.idx_score].value = score
        
        # Tags (Append column if needed)
        # Note: if row length is less than idx_tags, we need to append empty cells?
        # openpyxl handles assignment to sparse cells usually?
        # Actually iter_rows yields tuples of cells. We might need to access ws.cell()
        # row is a tuple of cells. We cannot append to it.
        # We need to find the Coordinate of the Tags cell.
        
        row_idx = row[0].row # Get 1-based row index
        
        self.ws.cell(row=row_idx, column=self.idx_score+1, value=score) # Column is 1-based
        # Tags saved via save_current_tags above

        
        # 3. Save Excel
        try:
            self.wb.save(self.excel_path)
        except Exception as e:
            print(f"Save Error: {e}")
            
        # 4. Move File to Score_X
        old_path = data["path"]
        if old_path and os.path.exists(old_path):
            target_dir = os.path.join(self.current_folder, f"Score_{score}")
            if not os.path.exists(target_dir): os.makedirs(target_dir)
            
            if os.path.dirname(old_path) != target_dir:
                 new_path = os.path.join(target_dir, fname)
                 try:
                     shutil.move(old_path, new_path)
                     data["path"] = new_path
                 except Exception as e:
                     print(f"Move Error: {e}")
        
        # 5. Update UI Tree
        self.lbl_current_score.configure(text=str(score))
        self.apply_score_color(score)
        
        item_id = data["item_id"]
        self.tree.set(item_id, "score", str(score))
        
        # Tags update
        cur_tags = []
        if score >= 8: cur_tags.append("high")
        elif score <= 3: cur_tags.append("low")
        else: cur_tags.append("mid")
        self.tree.item(item_id, tags=cur_tags)
        
        # 6. Next
        if self.auto_advance:
            self.select_next()
            
    def select_next(self):
        sel = self.tree.selection()
        if not sel: return
        next_id = self.tree.next(sel[0])
        if next_id:
            self.tree.selection_set(next_id)
            self.tree.focus(next_id)
            self.tree.see(next_id)

    def on_sync(self):
        # Sync Score_1 to Score_10 + HighScore/LowScore legacy
        # Check config
        config_path = "uploader_config.json"
        target_path = ""
        
        if os.path.exists(config_path):
            with open(config_path, "r") as f:
                c = json.load(f)
                target_path = c.get("upload_target", "")
        
        if not target_path or not os.path.exists(target_path):
            tk.messagebox.showinfo("Setup", "Please select the Sync Target folder.")
            target_path = filedialog.askdirectory(title="Select Sync Target")
            if not target_path: return
            with open(config_path, "w") as f:
                json.dump({"upload_target": target_path}, f)
                
        ans = tk.messagebox.askyesno("Sync", f"Upload assets to:\n{target_path}\nContinue?")
        if not ans: return
        
        count = 0
        
        # Loop main folders
        folders = [f"Score_{i}" for i in range(1, 11)] + ["HighScore", "LowScore"]
        
        for fd in folders:
            src_d = os.path.join(self.current_folder, fd)
            if os.path.exists(src_d):
                dst_d = os.path.join(target_path, fd)
                if not os.path.exists(dst_d): os.makedirs(dst_d)
                
                for f in os.listdir(src_d):
                     if f.lower().endswith(".wav"):
                         s_f = os.path.join(src_d, f)
                         d_f = os.path.join(dst_d, f)
                         if not os.path.exists(d_f):
                             shutil.copy2(s_f, d_f)
                             count += 1
                             
        # Manifest
        import socket
        pc_name = socket.gethostname()
        folder_name = os.path.basename(self.current_folder)
        man_name = f"{pc_name}_{folder_name}_manifest.xlsx"
        shutil.copy2(self.excel_path, os.path.join(target_path, man_name))
        
        tk.messagebox.showinfo("Sync Complete", f"Synced {count} new files.")

if __name__ == "__main__":
    import sys
    start_dir = None
    if len(sys.argv) > 1:
        start_dir = sys.argv[1]
        
    app = SFXReviewerApp(start_dir)
    app.mainloop()
