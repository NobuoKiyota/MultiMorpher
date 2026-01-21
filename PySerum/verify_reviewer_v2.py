
import os
import shutil
import openpyxl
import customtkinter as ctk # Need to mock this or init headless
from sfx_reviewer_app import SFXReviewerApp

# Setup
TEST_DIR = "Output/TestBatch"
if os.path.exists(TEST_DIR): shutil.rmtree(TEST_DIR)
os.makedirs(TEST_DIR)

# Dummy Wav
wav_path = os.path.join(TEST_DIR, "test.wav")
with open(wav_path, "w") as f: f.write("dummy audio")

# Dummy Excel
xls_path = os.path.join(TEST_DIR, "generation_log.xlsx")
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["Score", "File Name", "Param1", "Date", "Version"]) # Headers
ws.append([None, "test.wav", "0.5", "2026-01-01", "1.0.0"]) # Data
wb.save(xls_path)

print("Setup Complete.")

# Init App (Mocking Mainloop)
# CTk might try to open window. We hope it doesn't block logic execution.
app = SFXReviewerApp()
app.withdraw() # Hide window

# Load Batch
print("Loading Batch...")
app.load_batch(TEST_DIR)

# Verify Load
if "test.wav" not in app.data_map:
    print("FAIL: File not loaded")
    exit(1)

# Set Tag
print("Setting Tag...")
app.tag_vars["Noisy"] = ctk.BooleanVar(value=True) # Manually trigger checkbox var
app.current_file = "test.wav"

# Set Score (Trigger Logic)
print("Setting Score 8...")
app.set_score(8)

# Verify File Move
expected_path = os.path.join(TEST_DIR, "Score_8", "test.wav")
if os.path.exists(expected_path):
    print("PASS: File moved to Score_8")
else:
    print(f"FAIL: File not found at {expected_path}")
    print(f"Current location: {app.data_map['test.wav']['path']}")

# Verify Excel Update
print("Verifying Excel...")
wb2 = openpyxl.load_workbook(xls_path)
ws2 = wb2.active
rows = list(ws2.iter_rows(values_only=True))
headers = rows[0]
data = rows[1]

# Index Check
idx_score = headers.index("Score")
idx_tags = headers.index("Tags") # Should be appended

if data[idx_score] == 8:
    print("PASS: Score updated to 8")
else:
    print(f"FAIL: Score is {data[idx_score]}")

if data[idx_tags] == "Noisy":
    print("PASS: Tag is 'Noisy'")
else:
    print(f"FAIL: Tag is '{data[idx_tags]}'")

app.destroy()
print("Verification Complete.")
