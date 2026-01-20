import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import random
from pysfx_param_config import PySFXParams
from pysfx_color_config import PySFXColors

class ExcelConfigLoader:
    DEFAULT_FILENAME = "Factory_Parameters.xlsx"

    @staticmethod
    def get_excel_path(workspace_root=None):
        if workspace_root is None:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            return os.path.join(os.path.dirname(current_dir), ExcelConfigLoader.DEFAULT_FILENAME)
        return os.path.join(workspace_root, ExcelConfigLoader.DEFAULT_FILENAME)

    @staticmethod
    def _apply_style(ws, row_idx, col_count, group_id=None):
        """Applies borders and background color based on group_id."""
        thin_border = Border(left=Side(style='thin'), 
                               right=Side(style='thin'), 
                               top=Side(style='thin'), 
                               bottom=Side(style='thin'))
        
        hex_color = None
        if group_id:
            hex_color = PySFXColors.get_excel_color(group_id)
        
        # If hex_color is None/White/Transparent, skip fill
        fill = None
        if hex_color and hex_color.lower() not in ["ffffffff", "transparent", "none"]:
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            
        for i in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=i)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    @staticmethod
    def create_template_excel(path):
        """Creates a template Excel file with Factory, Effects, and Weights sheets."""
        # Note: If file exists, we generally don't overwrite user data, 
        # BUT user requested template update. We'll only create if missing 
        # OR if we want to add missing sheets? 
        # Current logic: Create if missing. If exists, maybe we create separate template?
        # User said "Use template regeneration to handle new sheets". 
        # If file exists, we can try to add sheets if missing.
        
        wb = None
        if os.path.exists(path):
            try:
                wb = openpyxl.load_workbook(path)
            except:
                wb = None
        
        if wb is None:
            wb = openpyxl.Workbook()
            # Remote default sheet
            if "Sheet" in wb.sheetnames:
                 del wb["Sheet"]
                 
        # --- Sheet 1: Factory ---
        if "Factory" not in wb.sheetnames:
            ws_fac = wb.create_sheet("Factory", 0)
            headers = ["Parameter Name", "Base Value", "Probability (%)", "Min Value", "Max Value", "Description"]
            ws_fac.append(headers)
            
            # Styling Header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
            for c in ws_fac[1]: 
                c.font = header_font
                c.fill = header_fill
            
            params = PySFXParams.get_sorted_params()
            for p in params:
                # Default Base Value from Param Config
                base_val = p.default if p.default is not None else 0
                
                ws_fac.append([p.name, base_val, 100, 
                               p.min if p.min is not None else 0, 
                               p.max if p.max is not None else 1, 
                               p.desc])
                ExcelConfigLoader._apply_style(ws_fac, ws_fac.max_row, 6, p.group)
                
            ws_fac.column_dimensions['A'].width = 25
            ws_fac.column_dimensions['F'].width = 50

        # --- Sheet 2: Effects ---
        if "Effects" not in wb.sheetnames:
            ws_fx = wb.create_sheet("Effects")
            headers = ["Module", "Parameter", "Base Value", "Probability (%)", "Min Value", "Max Value", "Description"]
            ws_fx.append(headers)
            
            # Header Style
            for c in ws_fx[1]: 
                c.font = Font(bold=True, color="FFFFFF"); 
                c.fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")

            # Default Effects Config (Full Param Structure)
            # Tuple: (Module, Param, Base, Prob, Min, Max, Desc)
            fx_data = [
                ("Slicer", "threshold_db", -60, 0, -80, -40, "無音判定閾値(dB)"),
                ("Slicer", "min_interval_ms", 200, 0, 50, 500, "無音区間最小長(ms)"),
                ("Slicer", "min_duration_ms", 500, 0, 100, 2000, "切り出し最小長(ms)"),
                ("Slicer", "pad_ms", 50, 0, 10, 200, "前後の余白(ms)"),
                
                ("Masker", "NoiseType", "Random", 0, 0, 0, "Noise Type (Random, Pink, etc)"),
                ("Masker", "MaskAmount_Min", 0.1, 0, 0.0, 0.5, "Masker強度 最小"),
                ("Masker", "MaskAmount_Max", 0.4, 0, 0.2, 0.8, "Masker強度 最大"),
                
                ("Normalizer", "target_time_min", 0.5, 0, 0.1, 3.0, "最終尺 最小(秒)"),
                ("Normalizer", "target_time_max", 3.0, 0, 1.0, 10.0, "最終尺 最大(秒)"),
            ]
            
            # Add simple mapping for color
            mod_color_map = {
                "Slicer": 1, # Grey
                "Masker": "Phaser", # Purple-ish
                "Normalizer": 11 # Teal
            }
            
            for row in fx_data:
                ws_fx.append(row)
                mod = row[0]
                gid = mod_color_map.get(mod, 0)
                ExcelConfigLoader._apply_style(ws_fx, ws_fx.max_row, 7, gid)
                
            ws_fx.column_dimensions['A'].width = 15
            ws_fx.column_dimensions['B'].width = 25
            ws_fx.column_dimensions['G'].width = 40

        # --- Sheet 3: Weights ---
        if "Weights" not in wb.sheetnames:
            ws_w = wb.create_sheet("Weights")
            headers = ["Route Name", "Weight (Prob)", "Description"]
            ws_w.append(headers)
            for c in ws_w[1]: 
                c.font = Font(bold=True, color="FFFFFF"); 
                c.fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")

            weights = [
                ("Transformer", 30, "Transformerのみ"),
                ("Masker", 30, "Maskerのみ"),
                ("Through", 20, "加工なし(Skip)"),
                ("TransMask", 10, "Trans -> Mask"),
                ("MaskTrans", 10, "Mask -> Trans")
            ]
            
            for row in weights:
                ws_w.append(row)
                ExcelConfigLoader._apply_style(ws_w, ws_w.max_row, 3, "Master") # Blueish
                
            ws_w.column_dimensions['A'].width = 20
            ws_w.column_dimensions['C'].width = 30
            
        try:
            wb.save(path)
            print(f"Update/Create Parameter Template: {path}")
        except Exception as e:
            print(f"Error creating template: {e}")

    @staticmethod
    def load_config(path):
        """
        Loads ALL configs.
        Returns: {
            "Factory": dict, 
            "Effects": dict, 
            "Weights": dict
        }
        """
        if not os.path.exists(path):
            ExcelConfigLoader.create_template_excel(path)
            return None # Force re-load or handle empty

        print(f"Loading Excel: {path}")
        result = {"Factory": {}, "Effects": {}, "Weights": {}}
        
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            
            # 1. Factory
            if "Factory" in wb.sheetnames:
                ws = wb["Factory"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    # A:Name, B:Base, C:Prob, D:Min, E:Max
                    p_name = str(row[0]).strip()
                    p_base = row[1] if row[1] is not None else 0
                    p_prob = float(row[2]) if row[2] is not None else 0
                    p_min = row[3] if row[3] is not None else 0
                    p_max = row[4] if row[4] is not None else 0
                    
                    result["Factory"][p_name] = {
                        "value": p_base,
                        "min": p_min, "max": p_max, "probability": p_prob,
                        "random": (p_prob >= 100)
                    }
                    
            # 2. Effects
            if "Effects" in wb.sheetnames:
                ws = wb["Effects"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    # Module(A), Param(B), Base(C), Prob(D), Min(E), Max(F)
                    mod = str(row[0]).strip()
                    param = str(row[1]).strip()
                    p_base = row[2] if row[2] is not None else 0
                    p_prob = float(row[3]) if row[3] is not None else 0
                    p_min = row[4] if row[4] is not None else 0
                    p_max = row[5] if row[5] is not None else 0
                    
                    if mod not in result["Effects"]: result["Effects"][mod] = {}
                    
                    # Store Full Structure
                    result["Effects"][mod][param] = {
                        "value": p_base,
                        "min": p_min, "max": p_max, "probability": p_prob,
                        "random": (p_prob >= 100)
                    }
                    
            # 3. Weights
            if "Weights" in wb.sheetnames:
                ws = wb["Weights"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]: continue
                    # Route(A), Weight(B)
                    route = str(row[0]).strip()
                    w = row[1] if row[1] is not None else 0
                    result["Weights"][route] = w
                    
            return result
            
        except Exception as e:
            print(f"Error loading Excel config: {e}")
            return None
