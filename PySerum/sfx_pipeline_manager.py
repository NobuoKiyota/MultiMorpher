import os
import shutil
import glob
import random
import time
import datetime
import traceback
import openpyxl
import re

# Import Engines
from pysfx_factory import PyQuartzFactory
from pysfx_transformer_engine_tracked import QuartzTransformerEngineTracked
from pysfx_masker_engine import QuartzMaskerEngine
from pysfx_slicer_engine import QuartzSlicerEngine
from pysfx_normalizer_engine import QuartzNormalizerEngine
from pysfx_logger import SFXLogger
from pysfx_excel_loader import ExcelConfigLoader

class SFXPipeline:
    def __init__(self, workspace_root=None):
        if workspace_root is None:
            workspace_root = os.path.dirname(os.path.abspath(__file__))
            
        self.root_dir = os.path.join(workspace_root, "Pipeline_Output")
        self.factory = PyQuartzFactory()
        self.transformer = QuartzTransformerEngineTracked()
        self.masker = QuartzMaskerEngine()
        self.slicer = QuartzSlicerEngine()
        self.normalizer = QuartzNormalizerEngine()
        
    def _resolve_fx_params(self, config_dict):
        """
        Resolves a dict of {Param: {val, prob, min, max}} into {Param: Value}.
        Handles randomization.
        """
        resolved = {}
        if not config_dict: return resolved
        
        for k, node in config_dict.items():
            # If node is simple value, use it
            if not isinstance(node, dict):
                resolved[k] = node
                continue
                
            # If node is Full Config
            if "value" in node:
                prob = node.get("probability", 0)
                if prob > 0 and (prob >= 100 or random.uniform(0, 100) < prob):
                    # Random
                    v_min = float(node.get("min", 0))
                    v_max = float(node.get("max", 0))
                    if v_min > v_max: v_min, v_max = v_max, v_min
                    resolved[k] = random.uniform(v_min, v_max)
                else:
                    # Base
                    resolved[k] = node["value"]
            else:
                # Fallback? Or maybe it's nested?
                # Assume if no 'value' key, it might be raw dict? 
                # For safety, if it looks like a param struct without value, ignore or 0? 
                # Should not happen with new loader.
                pass
                
        return resolved

    def run_pipeline(self, batch_name, total_count=50, source_count=None, factory_settings=None, routing_weights=None, excel_path=None):
        """
        Gacha Algorithm Pipeline (Advanced Loop).
        1. Create Asset Pool (Factory -> Slicer) using source_count.
        2. Draw from Pool and Process (Loop 3-7 times).
        3. Determine Route based on weights.
        """
        
        # Default Weights
        if not routing_weights:
            routing_weights = {"Transformer":30, "Masker":30, "Through":20, "TransMask":10, "MaskTrans":10}
            weighted_from_excel = False # Check if we should override later
        else:
            weighted_from_excel = False # Manual provided

        # Normalize Weights
        w_keys = list(routing_weights.keys())
        w_vals = list(routing_weights.values())
        
        # Setup Directories
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if not batch_name: batch_name = f"Batch_{timestamp}"
        
        base_dir = os.path.join(self.root_dir, batch_name)
        dir_pool_raw = os.path.join(base_dir, "00_Pool_Raw")
        dir_pool_sliced = os.path.join(base_dir, "00_Pool_Sliced")
        dir_temp = os.path.join(base_dir, "01_Temp_Processing") # For loop intermediates
        dir_final = os.path.join(base_dir, "05_Final_Normalized")
        
        for d in [dir_pool_raw, dir_pool_sliced, dir_temp, dir_final]:
            os.makedirs(d, exist_ok=True)
            
        print(f"=== SFX Pipeline Started: {batch_name} (Overview Target: {total_count}) ===")
        t_start_total = time.time()

        # --- Phase 1: Create Asset Pool ---
        if source_count is None: source_count = 5
        print(f"--- Phase 1: Generating Asset Pool (Source: {source_count}) ---")
        
        # 1. Factory (Load Excel Config)
        self.factory.out_dir = dir_pool_raw
        
        # Load Excel Logic
        if not excel_path:
             excel_path = ExcelConfigLoader.get_excel_path()
        
        full_config = ExcelConfigLoader.load_config(excel_path)
        
        factory_config_excel = full_config.get("Factory", {}) if full_config else {}
        effects_config = full_config.get("Effects", {}) if full_config else {}
        weights_config = full_config.get("Weights", {}) if full_config else {}
        
        # If no manual routing weights provided (None passed from UI when Excel checked), use Excel Weights
        if routing_weights is None or (len(w_keys) == 5 and w_vals[0]==30 and w_vals[2]==20): # Or simple check arg
             pass
        
        # Better: UI passes None if Excel checked.
        if routing_weights is None and weights_config:
             print("Using Routing Weights from Excel.")
             routing_weights = weights_config
             w_keys = list(routing_weights.keys())
             w_vals = list(routing_weights.values())
        elif routing_weights is None:
             # Default fallback
             routing_weights = {"Transformer":30, "Masker":30, "Through":20, "TransMask":10, "MaskTrans":10}
             w_keys = list(routing_weights.keys())
             w_vals = list(routing_weights.values())

        # Merge Factory Overrides
        
        factory_config_excel = full_config.get("Factory", {}) if full_config else {}
        effects_config = full_config.get("Effects", {}) if full_config else {}
        weights_config = full_config.get("Weights", {}) if full_config else {}
        
        # Merge Factory Overrides
        final_config = factory_config_excel if factory_config_excel else self.factory.get_random_config()
        
        if factory_settings:
            for k, v in factory_settings.items():
                if k in final_config: 
                    final_config[k]["value"] = v 
                    final_config[k]["random"] = False
        
        self.factory.run_advanced_batch(final_config, num_files=source_count)
        
        # 2. Slicer
        # Default Params
        s_params = {
            "threshold_db": -60,
            "min_interval_ms": 200,
            "min_duration_ms": 500, 
            "pad_ms": 50
        }
        # Override from Excel (Resolve Random ONCE for the batch slicing operation)
        if "Slicer" in effects_config:
            s_exc = self._resolve_fx_params(effects_config["Slicer"])
            s_params.update(s_exc)
            
        self.slicer.process_folder(dir_pool_raw, dir_pool_sliced, s_params, progress_cb=None)
        
        # Load Pool
        pool_files = glob.glob(os.path.join(dir_pool_sliced, "*.wav"))
        if not pool_files:
            print("Error: No assets in pool!")
            return
            
        print(f"Pool Created: {len(pool_files)} assets available.")
        
        # --- Load Factory Log for Traceability ---
        factory_log_path = os.path.join(dir_pool_raw, "generation_log.xlsx")
        factory_data_map = self._load_factory_log(factory_log_path)

        # --- Phase 2: Production Loop ---
        print(f"--- Phase 2: Production Loop (Target: {total_count}) ---")
        
        # Check existing files to determine start index
        start_index = 1
        existing_files = glob.glob(os.path.join(dir_final, "Gen_*_*.wav"))
        if existing_files:
            max_idx = 0
            for f in existing_files:
                # Expecting Gen_XXXX_...
                base = os.path.basename(f)
                try:
                    parts = base.split("_")
                    if len(parts) >= 2 and parts[1].isdigit():
                        idx = int(parts[1])
                        if idx > max_idx: max_idx = idx
                except: pass
            start_index = max_idx + 1
            print(f"Existing files detected. Starting sequence from #{start_index:04d}")
        
        generated_count = 0
        final_logs = []
        
        while generated_count < total_count:
            try:
                # 1. Draw from Pool
                src_file = random.choice(pool_files)
                src_name = os.path.basename(src_file)
                
                # 2. Determine Loop Count (Single Pass)
                loop_count = 1 
                
                # Prepare Temp Flow
                current_file = src_file
                history_routes = []
                processing_details = {} # Store runtime params here
                
                # Copy start file to temp
                # Temp working name doesn't matter much, but keep it unique per run
                work_file = os.path.join(dir_temp, f"work_{generated_count}.wav")
                shutil.copy2(current_file, work_file)
                current_file = work_file
                
                for i in range(loop_count):
                    # Determine Route
                    route = random.choices(w_keys, weights=w_vals, k=1)[0]
                    history_routes.append(f"[{route}]")
                    
                    next_file = os.path.join(dir_temp, f"work_{generated_count}_next.wav")
                    
                    # Apply Route
                    success = True
                    if route == "Through":
                        shutil.copy2(current_file, next_file)
                        processing_details["Route"] = "Through"
                        
                    elif route == "Transformer":
                        t_params = {"NoiseType": "Random", "MaskAmount_Min": 0.2, "MaskAmount_Max": 0.6}
                        
                        if "Masker" in effects_config:
                             m_exc = self._resolve_fx_params(effects_config["Masker"])
                             t_params.update(m_exc)
                        
                        self.masker.process_file(current_file, next_file, t_params)
                        processing_details.update({f"Trans_{k}": v for k, v in t_params.items()})
                        
                    elif route == "Masker":
                        m_params = {"NoiseType": "Random", "MaskAmount_Min": 0.1, "MaskAmount_Max": 0.3}
                        if "Masker" in effects_config:
                             m_exc = self._resolve_fx_params(effects_config["Masker"])
                             m_params.update(m_exc)
                        self.masker.process_file(current_file, next_file, m_params)
                        processing_details.update({f"Mask_{k}": v for k,v in m_params.items()})
                        
                    elif route == "TransMask" or route == "MaskTrans":
                        # Double pass
                        temp_mid = os.path.join(dir_temp, "mid.wav")
                        m_params = {"NoiseType": "Random"}
                        if "Masker" in effects_config: 
                             m_exc = self._resolve_fx_params(effects_config["Masker"])
                             m_params.update(m_exc)
                        
                        self.masker.process_file(current_file, temp_mid, m_params)
                        self.masker.process_file(temp_mid, next_file, m_params) 
                        processing_details.update({f"Double_{k}": v for k,v in m_params.items()})
                        
                    # Swap
                    if os.path.exists(next_file):
                        shutil.move(next_file, current_file)
                    else:
                        pass # Fail? Keep current.
                        
                # 3. Final Normalize
                # Use start_index + generated_count to ensure unique sequence across runs
                current_seq_id = start_index + generated_count
                final_name = f"Gen_{current_seq_id:04d}_L{loop_count}.wav"
                final_out = os.path.join(dir_final, final_name)
                n_params = {"target_time_min": 0.5, "target_time_max": 3.0} 
                if "Normalizer" in effects_config:
                    n_exc = self._resolve_fx_params(effects_config["Normalizer"])
                    n_params.update(n_exc)
                
                self.normalizer.process_single_file(current_file, final_out, n_params)
                processing_details.update({f"Norm_{k}": v for k,v in n_params.items()})
                
                # Log
                if os.path.exists(final_out):
                     log_entry = {
                         "File Name": final_name,
                         "Score": "",
                         "Route": " -> ".join(history_routes),
                         "Source_File": src_name
                     }
                     
                     # 1. Merge Processing Details
                     log_entry.update(processing_details)
                     
                     # 2. Merge Factory Data
                     # Try to match Source Name (remove _01 suffix)
                     # Pattern: RawName + _\d+ + .wav
                     # Example: Quartz_..._01.wav -> Quartz_... .wav (Logic from Slicer)
                     # Actually Factory Log key is filename (Quartz_... .wav).
                     # Sliced file is Quartz_..._01.wav
                     
                     raw_base_match = re.match(r"(.*)_\d+(\.[a-zA-Z0-9]+)$", src_name)
                     if raw_base_match:
                         raw_key = raw_base_match.group(1) + raw_base_match.group(2)
                     else:
                         raw_key = src_name # Maybe unsliced or full file
                         
                     if raw_key in factory_data_map:
                         # Merge factory params with prefix "F_"? Or direct?
                         # Direct seems cleaner but might collide. 
                         # User wants "Detailed Info", collision unlikely unless same param names.
                         # Factory params are like "Duration", "Voices". Processing are "MaskAmount".
                         # Safe to merge directly.
                         log_entry.update(factory_data_map[raw_key])
                         
                     final_logs.append(log_entry)
                     generated_count += 1
                     print(f"Generated {generated_count}/{total_count} (Route: {route})", end="\r")

            except Exception as e:
                print(f"Loop Error: {e}")
                traceback.print_exc()
                
        print("\n=== Pipeline Complete ===")
        
        # Manifest Logic: Append if exists
        manifest_path = os.path.join(dir_final, "final_manifest.xlsx")
        
        # Collect all unique keys for header
        all_keys = ["Score", "File Name", "Route", "Source_File"] # Order priority
        seen = set(all_keys)
        
        # Scan all logs to find extra keys
        for l in final_logs:
            for k in l.keys():
                if k not in seen:
                    all_keys.append(k)
                    seen.add(k)
        
        if os.path.exists(manifest_path):
            try:
                wb = openpyxl.load_workbook(manifest_path)
                ws = wb.active
                # If appending, we assume headers roughly match or we just append rows based on CURRENT keys.
                # Ideally we should map to existing headers.
                existing_headers = [c.value for c in ws[1]]
                
                # Update header list if new keys introduced?
                # For simplicity in this fix, we append new keys to the end of header row if missing
                # But safer to just map what we have to existing, and append extras.
                
                for k in all_keys:
                    if k not in existing_headers:
                        # Add new header column
                        ws.cell(row=1, column=len(existing_headers)+1, value=k)
                        existing_headers.append(k)
                
                # Now append data mapped to headers
                for l in final_logs:
                    row = []
                    for h in existing_headers:
                        row.append(l.get(h, ""))
                    ws.append(row)
                    
            except Exception as e:
                print(f"Warning: Failed to append manifest, creating new. Error: {e}")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(all_keys)
                for l in final_logs:
                    row = [l.get(k, "") for k in all_keys]
                    ws.append(row)
        else:
            # Create New
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(all_keys)
            for l in final_logs:
                row = [l.get(k, "") for k in all_keys]
                ws.append(row)
            
        try:
            wb.save(manifest_path)
        except Exception as e:
            print(f"Error Saving Manifest: {e}")
        
        # Performance Report (Append or Write?)
        # Report is simple txt, overwrite is fine or append. Let's append with timestamp.
        dur_total = time.time() - t_start_total
        with open(os.path.join(base_dir, "performance_report.txt"), "a") as f:
             f.write(f"Run {datetime.datetime.now()}: Generated {generated_count} files in {dur_total:.2f}s\n")
        print(f"Report saved to {base_dir}")

    def _load_factory_log(self, path):
        """Loads generation_log.xlsx into a dict: filename -> {param: val}"""
        data = {}
        if not os.path.exists(path): return data
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = [c.value for c in ws[1]] # Row 1
            
            if "File Name" not in headers: return data
            idx_fname = headers.index("File Name")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                fname = row[idx_fname]
                if not fname: continue
                
                entry = {}
                for i, val in enumerate(row):
                    header = headers[i]
                    if header in ["Score", "File Name", "Date"]: continue # Skip metadata
                    entry[header] = val
                data[fname] = entry
        except Exception as e:
            print(f"Warning: Could not load factory log: {e}")
            
        return data

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--count", type=int, default=10)
    parser.add_argument("--name", type=str, default="")
    args = parser.parse_args()
    
    pipeline = SFXPipeline()
    pipeline.run_pipeline(args.name, args.count)
